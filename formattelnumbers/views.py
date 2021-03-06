from django.shortcuts import render, redirect
from formattelnumbers import cleaningTelNum, cleaningTelNumPreparation
from apirelay import dataProcessing

#format uploaded files
#from openpyxl import *
import openpyxl
from django.http import HttpResponse

# Create your views here.

from .forms import format_tel_numbers_input_form, uploaded_documents_form
from .models import uploaded_documents

#turn tel numbers into a list
def turn_textdata_to_list(textdata):
    listTextData = textdata.split("\r\n")
    return listTextData

#turn list into rows
def turn_list_to_row(numberList):     
    rowTextData = "\r\n".join(str(x) for x in numberList)
    return rowTextData

# Create your views here.
def process_numbers(listTextData, telCountry = None):    
    listResult = []
    for number in listTextData:
        number = cleaningTelNumPreparation.fix_telephone_format(number, telCountry)
        listResult.append(number)
    return listResult

def format_tel_numbers_input(request):
    if request.method == 'POST':
        numbersToFormat = format_tel_numbers_input_form(request.POST)

        if numbersToFormat.is_valid():
            isCountyInList = dataProcessing.check_country_code(numbersToFormat['countryCode'].value())
            if isCountyInList == "non_existing_country":
                numbersToFormat = format_tel_numbers_input_form(initial={"inputText" : "This country code is not selectable."})
                return render(request, 'formatTelNumbers/formattelnumers.html', {'numbersToFormat': numbersToFormat})


            textdata = numbersToFormat['inputText'].value()
            countrydata = numbersToFormat['countryCode'].value()
            #turn tel numbers into a list
            listTextData = turn_textdata_to_list(textdata)
            
            #processing of telephone numbers
            listResult = process_numbers(listTextData, countrydata)

            #turn list into rows
            rowTextData = turn_list_to_row(listResult)

            #return processed data
            data = {'inputText': rowTextData, "countryCode" : numbersToFormat['countryCode'].value()}
            numbersToFormat = format_tel_numbers_input_form(initial=data)

            return render(request, 'formatTelNumbers/formattelnumers.html', {'numbersToFormat': numbersToFormat})

    else:
        numbersToFormat = format_tel_numbers_input_form()
        return render(request, 'formatTelNumbers/formattelnumers.html', {'numbersToFormat': numbersToFormat})





def format_tel_numbers_upload(request):
    if request.user.is_authenticated:
        
        if request.method == 'POST':
            form = uploaded_documents_form(request.POST, request.FILES)
            if form.is_valid():

                documentName = form['document'].value()

                isFileSafe = cleaningTelNumPreparation.checkIfFileIsSafe(documentName)
                #isFileSafe = 'safe_to_work'

                if (isFileSafe == 'not_safe_to_work'):
                    form = uploaded_documents_form()

                elif (isFileSafe == 'safe_to_work'):        
                    
                    #expaning and then saving form var.
                    newForm = form.save(commit=False) #save obj but not commit
                    newForm.document_user = request.user #save user id into document_user expanded form
                    newForm.save() #same as form.save()

                    theFile = openpyxl.load_workbook("media/format_tel_number/" + str(documentName))
                    allSheetNames = theFile.sheetnames

                    for sheet in allSheetNames:
                        currentSheet = theFile[sheet]
                        specificCellLetter = (cleaningTelNumPreparation.find_specific_cell(currentSheet))
                        letter = cleaningTelNumPreparation.get_column_letter(specificCellLetter)
                        cleaningTelNumPreparation.get_all_values_by_cell_letter(letter, currentSheet)

                    #saving a file
                    theFile.save("media/format_tel_number/" + str(documentName))
                
                    #Downloading a file
                    downloadedFile = open("media/format_tel_number/" + str(documentName), 'rb')
                    response = HttpResponse(content=downloadedFile)
                    response['Content-Type'] = 'xlsx'
                    response['Content-Disposition'] = 'attachment; filename="%s"' \
                                                    % (documentName) #public name

                    import os
                    from django.conf import settings
                    os.remove(os.path.join(settings.MEDIA_ROOT, "format_tel_number/" + str(documentName)))                                
                    return response
                else:
                    form = uploaded_documents_form()
        else:
            form = uploaded_documents_form()
        return render(request, 'formatTelNumbers/format_tel_numbers_upload.html', {
            'form': form
        })
    else:
        return render(request,'index.html')


def render_formatTelNumbersUpload_description_page(request):
    return render(
    request,
    'toolDescriptionPages/formattelnumuploaddesc.html'
)